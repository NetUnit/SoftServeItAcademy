import copy
from os import write
import re
# pyxl
import openpyxl as open
from pathlib import Path
from openpyxl.cell.cell import MergedCell, Cell
# python
import string
import datetime
#import xlsxwriter
path_main = '/media/netunit/storage/SoftServeItAcademy/airport_petty_algorithms/carrier_register_li'
path_reserve = '/home/rostyslav/Общедоступные/temp_andrii/SoftServeItAcademy/airport_petty_algorithms/carrier_register_li'


class ParseXLSXData:

    '''
    ===========================================================
    This class represents the manufacturer of a certain product
    ===========================================================
    Attrs:
    :param path_main: ...
    :type name: str
    :param path_reserve: ...
    :type country: str
    :param registry_rt: ...
    :type date: str
    :param registry_rt: ...
    :type date: str
    .. note:: 
        Put some notes here...
    '''
    fuel_choices = ['RT', 'Jet A-1']
    path_main = '/media/netunit/storage/SoftServeItAcademy/airport_petty_algorithms/carrier_register_li'
    path_reserve = '/home/rostyslav/Общедоступные/temp_andrii/SoftServeItAcademy/airport_petty_algorithms/carrier_register_li'
    
    # RT:
    registry_rt = 'Реєстр Укртатнафта РТ.xlsx'
    report_rt = 'Звіт РТ.xlsx'

    # Jet A-1:
    registry_jet = 'Реєстр Укртатнафта Jet A-1.xlsx'
    report_jet = 'Звіт Jet A-1.xlsx'
    # def __init__(self, path_main=path_main, path_reserve=path_reserve):
    #     self.path_main=path_main
    #     self.path_reserve=path_reserve

    def select_fuel_type(self):
        select_fuel = input('Please select type of fuel U want to make a report of: ')


    def load_RT_file_initial(self, registry=registry_rt):
        '''
            accessing & loading the initial RT fuel data \n\
            file from already created MS Excel file
            
            :returns: active MS Excel sheet for registry RT fuel
        '''
        reestr_xlsx_file = Path(
            path_reserve,
            registry
            )
        reestr_obj = open.load_workbook(reestr_xlsx_file)
        reestr_sheet = reestr_obj.active
        return reestr_sheet
        return print('it\'s ok')

    def load_RT_file_final(self, report=report_rt):
        '''
            accessing & loading final RT fuel data \n\
            file from already created MS Excel file
            
            :returns: active MS Excel sheet for report RT fuel
        '''
        oblik_xlsx_file = Path(
            path_reserve,
            report
            )
        
        oblik_obj = open.load_workbook(oblik_xlsx_file)
        oblik_sheet = oblik_obj.active
        return oblik_sheet  

        return print('it\'s ok')

    def load_JET_file_initial(self, registry=registry_jet):
        '''
            accessing & loading the initial JET fuel data \n\
            file from already created MS Excel file
            
            :returns: active MS Excel sheet for registry JET fuel
        '''
        reestr_xlsx_file = Path(
            path_reserve,
            registry
            )
        reestr_obj = open.load_workbook(reestr_xlsx_file)
        reestr_sheet = reestr_obj.active
        return reestr_sheet
        return print('it\'s ok')

    def load_JET_file_final(self, report=report_jet):
        '''
            accessing & loading final Jet fuel data \n\
            file from already created MS Excel file
            
            :returns: active MS Excel sheet for report JET fuel
 
        '''
        oblik_xlsx_file = Path(
            path_reserve,
            report
            )
        
        oblik_obj = open.load_workbook(oblik_xlsx_file)
        oblik_sheet = oblik_obj.active
        return oblik_sheet  
        return print('it\'s ok')
        
    def clean_fields(self):
        oblik_sheet = self.load_JET_file_final()
        ## clean fields from a previous day report
        for i in range(1, len(oblik_sheet['E'])+1):
            cell = oblik_sheet[f'E{i}']
            condition = isinstance(cell, Cell)
            if condition:
                cell.value = None
            else:
                continue
        
        oblik_obj.save(report)



if __name__ == "__main__":
    instance = ParseXLSXData()
    instance.load_RT_file_initial()
    instance.load_RT_file_final()
    instance.load_JET_file_initial()
    instance.load_JET_file_final()


        # '''
        #     loading initial RT fuel data via openpyxl module from already created MS Excel file
        #     :param name: Describes the comapny name
        #     :type: str max_length = 40
        #     :param country: Depicts the manufacturer's country of origin
        #     :type: str max_length = 20
        #     :param year: depicts the foundation year of a company
        #     :type date: datetime autofiled
        #     :param image: input field for upload/download of images
        #     :type image: 
        #     :param file: input field for upload/download of files
 
        # '''

# def parse_string(*args, **kwargs):
#     '''
#         We will be using all retrieved data from xlsx file
#         as local variables. After slicing the data we would then ask 
#         a user how to save fields or rewrite data to the same/another
#         file. 

#         NOTE: make a script executable file then 
#     '''

#     ## convert and find xlsx file --> make a widget for file location later
#     ## [.....] ADD WIDGET HERE

#     ## Jet A-1 
#     '''
#     registry = 'Реєстр Укртатнафта Jet A-1.xlsx'
#     reestr_lsx_file = Path(
#         path_reserve,
#         'Реєстр Укртатнафта Jet A-1.xlsx'
#         )   
#     '''

#     ## RT 
#     registry = 'Реєстр Укртатнафта РТ.xlsx'
#     reestr_lsx_file = Path(
#         path_reserve,
#         'Реєстр Укртатнафта РТ.xlsx'
#         )

#     reestr_obj = open.load_workbook(reestr_lsx_file)
    
#     '''
#     # Jet A-1
#     report = 'Звіт Jet A-1.xlsx'
#     oblik_lsx_file = Path(
#         path_reserve,
#         'Звіт Jet A-1.xlsx'
#         )
#     '''

#     ## RT
#     report = 'Звіт РТ.xlsx'
#     oblik_lsx_file = Path(
#         path_reserve,
#         report
#         )

#     oblik_obj = open.load_workbook(oblik_lsx_file)

#     # print(type(oblik_obj))

#     reestr_sheet = reestr_obj.active
#     oblik_sheet = oblik_obj.active

#     ## clean fields from a previous day report
#     for i in range(1, len(oblik_sheet['E'])+1):
#         cell = oblik_sheet[f'E{i}']
#         condition = isinstance(cell, Cell)
#         if condition:
#             cell.value = None
#         else:
#             continue
    
#     oblik_obj.save(report)

#     ## convert lambda function for date convert- ---> allows to convert date's if sheet was changed
#     ## NOTE: remake this function after beacuse if the cell is not a timedatae format it'll raise an error
#     ## to an appropriate format: '%Y-%m-%d'
#     converted = lambda date: datetime.date.strftime(date, '%Y-%m-%d')
    
    
#     # get selected date in sstr format
#     date = input('Please select the date of report, use DD-MM-YYYY format: ')
#     format = '%d-%m-%Y'
#     convert_to_dt = lambda date: datetime.datetime.strptime(date, format)
#     date = convert_to_dt(date)
    
#     # from datetime to date 
#     date = date.date() ## <class 'datetime.date'>
#     # print(dt)
#     # print(type(dt))
    
#     # get today's date
#     today = datetime.date.today() ## 2021-07-24 <class 'datetime.date'>
    
#     # print(today)
#     # print(type(today))
#     # print(dt==today)

#     ## create a dictionary which allows duplicates keys + nested values
#     #  --> remake to a the LC later
#     items = dict([])

#     for i in range(1, len(reestr_sheet['L'])+1):
#         condition = reestr_sheet[f'L{i}'].value != None
#         if condition:
#             items[reestr_sheet[f'L{i}'].value.lower()] = []

#     ## iterate through reestr sheet cells and get filter by todays date
#     for i in range(1, len(reestr_sheet['A'])+1):
        
#         ## avoid errors when iterating blanc fields
#         condition = reestr_sheet[f'A{i}'].value != None

#         ## use this condition to raise date_error and render to a window of errors
#         ## whnen incorrect date was put or wrong format 
#         condition2 = isinstance(reestr_sheet[f'A{i}'].value, datetime.date) 
        
#         if condition:
#             filter_by_date_cells = list(filter(
#                 # change the date to today if U want to get automatic yesterday report
#                 lambda item: str(item.value)[:10] == converted(date), reestr_sheet[i] 
#                 ))
        
#             if any(filter_by_date_cells):
                
#                 ## this condition will avoid adding Nones to nested dict values
#                 ## avoid: TypeError: unsupported operand type(s) for +: 'int' and 'NoneType'
#                 not_nones_cells = reestr_sheet[f'L{i}'].value != None and reestr_sheet[f'F{i}'].value != None
#                 items[reestr_sheet[f'L{i}'].value.lower()].append(reestr_sheet[f'F{i}'].value) if not_nones_cells else 0
    
#     print(items)
    
#     ## getting collected amount values from sheet reestr by carrier
#     ## returns dict with the next data: key - carrier, value - daily amount in kgs
#     def add_amounts(items):
#         daily_amount = 0
#         for key, value in items.items():
#             items[key] = sum(value)
#             daily_amount += sum(value)
#         return items

#     items = add_amounts(items)
#     print(items)

#     ## writing fields E5-E22 using prapared dict of items
#     ## with 'Zvit' Fields C5-C22 as keys
#     ## stage2: write matched by key values into cells through
#     ## get from loop indices - i
#     try:

#         i = 0
#         for carrier_cell in oblik_sheet[f'C']:

#             ## condition willl avoid exceptions
#             ##  when using blank cell as a key 
#             condition4 = carrier_cell.value is not None
#             record_value = items.get(str(carrier_cell.value).lower()) if condition4 else 0
            
#             ## iteration should be in this block 
#             ## as we need to get all indices of cells and sort only those
#             ## where carrier/key - value from items were found
#             i += 1
#             if record_value:
#                 oblik_sheet[f'E{i}'].value = record_value            
#                 oblik_obj.save(report) # change this field to automatic     

#     except:
#         print('Render a possible error here to the window')

    
#     ## input fuel arrival:
#     try:
#         fuel_arrived = int(input('Введіть надходження палива, кг: '))
#         i = 0
#         for cell in oblik_sheet[f'C']:
#             i += 1
#             pattern = re.compile('Прибуток палива')
#             ## find appropriate cell condition to write daily fuel suppply 
#             result = pattern.match(str(cell.value)) is not None
#             if result:
#                 oblik_sheet[f'E{i}'].value = fuel_arrived 
#                 oblik_obj.save(report) # change this field to automatic

#     except ValueError as input_err:
#         print('Please do not enter letters')
#         pass

#     except (TypeError, AttributeError) as input_error:
#         print('Please enter a correct amount')
#         pass


    
#     ## date of he report header B3
#     try:
#         ## block for automatic yesterday report
#         # automatic date report 
#         # yest = datetime.date.today() - datetime.timedelta(days=1)
#         # yest = datetime.date.strftime(yest, '%d.%m.%Y')
#         # today = datetime.date.strftime(today, '%d.%m.%Y')

#         ## variable value will be added for todays residue calculation
#         # oblik_sheet[f'B{3}'].value = f'Дата: {today}'
#         # oblik_obj.save(report) ## change this field to automatic
#         # #print(fuel_residue) if condition else 'Nothing'
        
#         ## block for preselected date of the report
#         ## date for prior of this report date
#         prior_rep_date = date - datetime.timedelta(days=1)
#         prior_rep_date = datetime.date.strftime(prior_rep_date, '%d.%m.%Y')
#         date = datetime.date.strftime(date, '%d.%m.%Y')
        
#         oblik_sheet[f'B{3}'].value = f'Дата: {date}'
#         oblik_obj.save(report) ## change this field to automatic
        
#     except (TypeError, ValueError) as input_error:
#         print(input_error)
#         pass

#     ## input fuel residue:
#     ## up-to-date fuel residue calculation (at the end of a current date 23:59:59)
#     try:

#         fuel_residue = int(input(f'Введіть залишок палива КГ, на дату {prior_rep_date} 23:59: '))
#         ## avoid errors when blanc field of fuel supply (no fuel supply daily)
#         condition = fuel_residue is not None
#         # E23
#         daily_amount = sum([i for i in items.values()])
        
#         # E26
#         residue_today = fuel_residue + fuel_arrived - daily_amount
#         pattern1 = re.compile('Всього видано на пероні')
#         pattern2 = re.compile('Залишок на складі ПММ')

#         i = 0
#         for cell in oblik_sheet[f'C']:
#             i += 1
#             search1 = pattern1.match(str(cell.value)) is not None
#             search2 = pattern2.match(str(cell.value)) is not None

#             if search1:
#                 ### FIELD E23 field
#                 oblik_sheet[f'E{i}'].value = daily_amount
#                 oblik_obj.save(report) # change this field to automatic

#             if search2:
#                 ### FIELD E26 field
#                 oblik_sheet[f'E{i}'].value = residue_today
#                 oblik_obj.save(report) # change this field to automatic
#     except Exception as err:
#         print(err)
#         pass 

            
# parse_string()