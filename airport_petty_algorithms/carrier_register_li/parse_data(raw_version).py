import copy
from os import write
import re
# pyxl
import openpyxl as open
from pathlib import Path
from openpyxl.cell.cell import MergedCell, Cell
# python
import string
import datetime
#import xlsxwriter
path_main = '/media/netunit/storage/SoftServeItAcademy/airport_petty_algorithms/carrier_register_li'
path_reserve = '/home/rostyslav/Общедоступные/temp_andrii/SoftServeItAcademy/airport_petty_algorithms/carrier_register_li'


class FuelArrivedField:
    
    msg = 'Please do not enter letters'
    msg2 = 'Please enter a correct amount'
    fuel_arrived = None

    def __init__(self):
        try:
            self.fuel_arrived = int(input('Enter the fuel supply, kg: '))
        
        except ValueError as err:
            print('Please do not enter letters')

        except (TypeError, AttributeError) as input_error:
            print(msg2)
        
        finally:
            exit() if self.fuel_arrived is None else True

class FuelResidueField(FuelArrivedField):
    
    fuel_residue = None
    def __init__(self, prev_rep_date):
        try:
            self.fuel_residue = int(
                input(f'Введіть залишок палива КГ ' + 
                f'на дату {prev_rep_date} 23:59: ')
            )
        except ValueError as err:
            print('Please do not enter letters')

        except (TypeError, AttributeError) as input_error:
            print(msg2)
        
        finally:
            exit() if self.fuel_residue is None else True


class ParseXLSXData:

    '''
    ===========================================================
    This class represents the manufacturer of a certain product
    ===========================================================
    Attrs:
    :param path_main: ...
    :type name: str
    :param path_reserve: ...
    :type country: str
    :param registry_rt: ...
    :type date: str
    :param registry_rt: ...
    :type date: str
    .. note:: 
        Put some notes here...
    '''
    fuel_choices = ['RT', 'Jet A-1']
    path_main = '/media/netunit/storage/SoftServeItAcademy/airport_petty_algorithms/carrier_register_li'
    path_reserve = '/home/rostyslav/Общедоступные/temp_andrii/SoftServeItAcademy/airport_petty_algorithms/carrier_register_li'
    
    # RT:
    registry_rt = 'Реєстр Укртатнафта РТ.xlsx'
    report_rt = 'Звіт РТ.xlsx'

    # Jet A-1:
    registry_jet = 'Реєстр Укртатнафта Jet A-1.xlsx'
    report_jet = 'Звіт Jet A-1.xlsx'

    select_fuel = input('Please select type of fuel U want to make a report of: ')
    
    # delete all this
    format = '%d-%m-%Y'
    format2 = '%Y-%m-%d'
    format3 = '%d.%m.%Y'

    # from <str> to <'datetime.datetime'>
    to_datetime =  lambda date, format: datetime.datetime.strptime(date, format) ## +++
    
    # from <'datetime.datetime'> to <'datetime.dt'>
    # date =  datetime.datetime.date()

    # from <str> to <'datetime.date'>
    to_date = lambda date, format: datetime.date.strftime(date, format) ### +++

    # from 'datetime.datetime'> to str
    to_str = lambda date, format: datetime.date.strftime(date, format) ## +++
    

    # fuel_arrived = 'Some data from UI' # 'Put it afterwrds' # get it from fuel_residue() & change it from input
    

    def __init__(self,
            report_rt=report_rt, registry_rt=registry_rt,
            report_jet=report_jet, registry_jet=registry_jet,
            select_fuel=select_fuel, fuel_choices=fuel_choices,
            format=format, format3 = format3
        ):
        ## init path here
        ## self.path = path_main ###---> select path from widget
        self.path = path_reserve
        ## init report here
        ## check if name of the file equals to preselected fuel which is report_jet

        ### init 

        # initialize report depending on type of fuel
        self.fuel_choices = fuel_choices
        self.select_fuel = select_fuel
        self.items = dict([])
        # correct input conditions
        fuel_selected = select_fuel in fuel_choices
        blanc_input = len(select_fuel) < 1 or None

        try:
            self.registry = None
            self.report = None

            if blanc_input:
                raise TypeError

            if not fuel_selected:
                raise ValueError

            self.registry = registry_rt if select_fuel=='RT' else registry_jet
            self.report = report_rt if select_fuel=='RT' else report_jet
            self.fuel_arrived = fuel_arrived

        except TypeError:
            print('Fuel selection shoudn\'t be blanc!')
        except ValueError:
            print('The fuel U\'ve  selected are not in the list!')
        except Exception as err3:
            print('Something went wrong. Try to relaunch the application')
        finally:
            exit() if self.registry is None or self.report is None else True
        
        print(self.registry, self.report + ' - This is ok: registry & report')

        # initialize date of report
        # this date is converted to satisfy needs of LibreOffice date format
        try:

                # 2021-08-24 00:00:00
                # <class 'datetime.datetime'>

            self.date = None
            date = input('Please select the date of report, use DD-MM-YYYY format: ')

            # correct date conditions
            blanc_input = len(date) < 1 or None

            if blanc_input:
                raise TypeError
            
            # from <str> to <'datetime.datetime'>
            date = ParseXLSXData.to_datetime(date, self.format)
            # date2 = ParseXLSXData.to_date(date, self.format)
            # date3 = datetime.datetime.today()
            # date3 = date3.date()
            # print(f'{date} djskjd')
            # print(f'{date2} djskjd')
            # print(f'{date3} djskjd')
            # from <str> to <'datetime.datetime'>
            # convert_to_datetime = lambda date: datetime.datetime.strptime(date, self.format)
            # date = convert_to_datetime(date) ## <class 'datetime.datetime'>

            # from datetime to date 
            # date = date.date() ## <class 'datetime.date'>
            
            # get today's date
            today = datetime.datetime.today() ## 2021-07-24 <class 'datetime.date'>
            

            date_in_range = date < today
            if not date_in_range:
                raise AttributeError
            
            self.date = date
            self.today = today

        except TypeError:
            print('Date selection shoudn\'t be blanc!')
        except ValueError:
            print('Please enter date in the correct format: yyyy-MM-dd')
        except AttributeError:
            print('Date is out of range')
        except Exception as err3:
            print('Something went wrong. Try to relaunch the application')
        finally:
            exit() if self.date is None else True
        
        print(f'{self.date} - This is ok: date')
        
    # def determine_reports(self):

    #     fuel_selected = self.select_fuel in self.fuel_choices
    #     if not fuel_selected:
    #         raise ValueError('The fuel U\'ve  selected are not in the list!')
    #     if self.select_fuel == 'RT':
    #         return self.registry_rt, self.report_rt
    #     if self.select_fuel == 'Jet A-1':
    #         return self.registry_jet, self.report_jet
    #     else:
    #         raise TypeError('Something went wrong. Check if fuel selection was wright')
    
 
    def load_file_initial(self):
        #reports = self.determine_reports()
        #registry = reports[0]
        
        print(self.registry)
        '''
            accessing & loading the initial RT fuel data \n\
            file from already created MS Excel file
            
            :returns: active MS Excel sheet for registry RT fuel
        '''
        reestr_xlsx_file = Path(
            self.path,
            self.registry
            )
        reestr_obj = open.load_workbook(reestr_xlsx_file)
        return reestr_obj
        # return print('it\'s ok')

    def activate_sheet_initial(self):
        '''
            Activate the initial jet fuel data sheet
            :returns: active MS Excel sheet for registry RT fuel
        '''
        registry_sheet = self.load_file_initial().active
        return registry_sheet

    @staticmethod
    def _is_method():
        return print('Ira')

    def load_file_final(self):

        # reports = self.determine_reports()
        # report = reports[1]

        print(self.report)

        '''
            accessing & loading final Jet fuel data \n\
            file from already created MS Excel file
            
            :returns: active MS Excel sheet for report JET fuel
 
        '''

        oblik_xlsx_file = Path(
            self.path,
            self.report
            )
            
        oblik_obj = open.load_workbook(oblik_xlsx_file)
        return oblik_obj
        # return print('it\'s ok')

    # def activate_sheet_final(self, oblik_obj):
    #     report_sheet = self.load_file_final().active
    # or ## report_sheet = report_sheet.active
    #     return report_sheet

    def check_cells(self):
        registry_sheet = self.activate_sheet_initial()
        
        a = set(i for i in 'osagau chemicals')
        b = set(i for i in 'osagau сhemicals')
        print(a, b)
        
        ## check fields for mistakes
        for i in range(1, len(registry_sheet['L'])+1):
            condition = registry_sheet[f'L{i}'].value != None
            if not condition:
                continue
            cell = registry_sheet[f'L{i}']
            ### add for each letters
            mistakes = [f'mistake #{i}:, cell: {cell}, symbol {cell.value}' for value in cell.value if not cell.value in string.ascii_letters] 
            print(mistakes)

    def clean_fields(self):
        oblik_obj = self.load_file_final()
        oblik_sheet = oblik_obj.active

        ## clean fields from a previous day report
        for i in range(1, len(oblik_sheet['E'])+1):
            cell = oblik_sheet[f'E{i}']
            condition = isinstance(cell, Cell)
            if condition:
                ## print(cell.value) ++
                cell.value = None
            else:
                continue
        
        # def get_determine_report
        oblik_obj.save(self.report)
        print('It\'s ok')

    def read_cells(self):
        print(self.date)
        ## create a dictionary which allows duplicates keys + nested values
        #  --> remake to a the LC later
        # reestr_sheet = registry_sheet
        registry_sheet = self.activate_sheet_initial()

        # assighn nested cells
        for i in range(1, len(registry_sheet['L'])+1):
            condition = registry_sheet[f'L{i}'].value != None
            if condition:
                self.items[registry_sheet[f'L{i}'].value.lower()] = []

        ## iterate through reestr sheet cells and get filter by todays date
        for i in range(1, len(registry_sheet['A'])+1):
            
            # print(type(registry_sheet[f'A{i}'].value)) # <class 'datetime.datetime'>

            ## avoid errors when iterating blanc fields
            condition = registry_sheet[f'A{i}'].value != None

            ## use this condition to raise date_error and render to a window of errors
            ## whnen incorrect date was put or wrong format 
            ## NEEDLESS
            condition2 = isinstance(registry_sheet[f'A{i}'].value, datetime.datetime) 
            # print(condition2) ## True
            
            if condition:
                # print(type(registry_sheet[f'A{i}'].value))
                filter_by_date_cells = list(filter(
                    # change the date to today if U want to get automatic yesterday report
                    
                    lambda item: item.value == self.date, registry_sheet[i] 
                    )
                )
                
                if any(filter_by_date_cells):
                    
                    ## this condition will avoid adding Nones to nested dict values
                    ## avoid: TypeError: unsupported operand type(s) for +: 'int' and 'NoneType'
                    not_nones_cells = registry_sheet[f'L{i}'].value != None and registry_sheet[f'F{i}'].value != None
                    self.items[registry_sheet[f'L{i}'].value.lower()].append(registry_sheet[f'F{i}'].value) if not_nones_cells else 0
        
        # print(self.items)
        return self.items
        
        ## getting collected amount values from sheet reestr by carrier
        ## returns dict with the next data: key - carrier, value - daily amount in kgs
    def add_amounts(self):
        daily_amount = 0
        for key, value in self.items.items():
            self.items[key] = sum(value)
            daily_amount += sum(value)
        return self.items


    #     ## writing fields E5-E22 using prapared dict of items
#     ## with 'Zvit' Fields C5-C22 as keys
#     ## stage2: write matched by key values into cells through
#     ## get from loop indices - i
    def write_cells(self):
        print(self.items) ## put add_amounts here instead of "__main__"
        oblik_obj = self.load_file_final()
        oblik_sheet = oblik_obj.active

        i = 0
        for carrier_cell in oblik_sheet[f'C']:

            ## condition willl avoid exceptions
            ##  when using blank cell as a key 
            condition4 = carrier_cell.value is not None
            # print(self.items.get(str(carrier_cell.value).lower()))
            record_value = self.items.get(str(carrier_cell.value).lower()) if condition4 else 0
            
            ## iteration should be in this block 
            ## as we need to get all indices of cells and sort only those
            ## where carrier/key - value from items were found
            i += 1
            if record_value:
                oblik_sheet[f'E{i}'].value = record_value            
                oblik_obj.save(self.report) # change this field to automatic     

    def input_fuel_arrival(self):
        # input fuel arrival:
        oblik_obj = self.load_file_final()
        oblik_sheet = oblik_obj.active

        # try:
        
        fuel_arrived_obj = FuelArrivedField()
        fuel_arrived = fuel_arrived_obj.fuel_arrived
        print(f'{fuel_arrived} + {type(fuel_arrived)}')
        # fuel_arrived = int(input('Enter the fuel supply, kg: '))
        
        # through the setattr
        setattr(self, 'fuel_arrived', fuel_arrived)
        print(f'{self.fuel_arrived}' + 'Bla bla bla. Ira\'s advice')

        i = 0
        for cell in oblik_sheet[f'C']:
            i += 1
            pattern = re.compile('Прибуток палива')
            ## find appropriate cell condition to write daily fuel suppply 
            result = pattern.match(str(cell.value)) is not None
            if result:
                oblik_sheet[f'E{i}'].value = fuel_arrived 
                oblik_obj.save(self.report) # change this field to automatic
        
        # except ValueError as input_err:
        #     print('Please do not enter letters')
        
        # except (TypeError, AttributeError) as input_error:
        #     print('Please enter a correct amount')
        

    ### write the date into a date-cell header
    def date_report_record(self):
        oblik_obj = self.load_file_final()
        oblik_sheet = oblik_obj.active
        
        # from 'datetime.datetime'> to str
        date = ParseXLSXData.to_str(self.date, self.format3)
        
        # date of he report header B3
        try:
            # print(type(self.date))
            ## block for automatic yesterday report
            # automatic date report 
            # yest = datetime.date.today() - datetime.timedelta(days=1)
            # yest = datetime.date.strftime(yest, '%d.%m.%Y')
            # today = datetime.date.strftime(today, '%d.%m.%Y')

            ## variable value will be added for todays residue calculation
            # oblik_sheet[f'B{3}'].value = f'Дата: {today}'
            # oblik_obj.save(report) ## change this field to automatic
            # #print(fuel_residue) if condition else 'Nothing'
            
            date = datetime.date.strftime(self.date, '%d.%m.%Y')
            oblik_sheet[f'B{3}'].value = f'Дата: {date}'
            oblik_obj.save(self.report) ## change this field to automatic
            
        except (TypeError, ValueError) as input_error:
            print(input_error)
    
    # get date of previous report 
    def previous_report_date(self):
        ## get data of previous to this report 
        prev_rep_date = self.date - datetime.timedelta(days=1)
        prev_rep_date = datetime.date.strftime(prev_rep_date, self.format3)
        return prev_rep_date

    # get fuel daily total amount in kg's
    def get_daily_amount(self):
        daily_amount = sum([i for i in self.items.values()])
        return daily_amount

    def get_fuel_residue(self):


        oblik_obj = self.load_file_final()
        oblik_sheet = oblik_obj.active

        ## input fuel residue:
        ## up-to-date fuel residue calculation (at the end of a current date 23:59:59)
        
        # fuel_residue = int(input(f'Enter the amount of fuel for the previous day, date: {prior_rep_date} 23:59: '))
        
        prev_rep_date = self.previous_report_date()
        # fuel_residue = int(input(f'Введіть залишок палива КГ, на дату {prev_rep_date} 23:59: '))

        fuel_residue_obj = FuelResidueField(prev_rep_date)
        fuel_residue = fuel_residue_obj.fuel_residue

        setattr(self, 'fuel_residue', fuel_residue)
        print(self.fuel_residue)
        print(type(self.fuel_residue))

        ## avoid errors when blanc field of fuel supply (no fuel supply daily)
        condition = fuel_residue is not None
        
        # E23
        daily_amount = self.get_daily_amount()
        
        # E26
        residue_today = self.fuel_residue + self.fuel_arrived - daily_amount

        
        print(type(self.fuel_arrived))

        pattern1 = re.compile('Всього видано на пероні')
        pattern2 = re.compile('Залишок на складі ПММ')
        
        i = 0
        for cell in oblik_sheet[f'C']:
            i += 1
            search1 = pattern1.match(str(cell.value)) is not None
            search2 = pattern2.match(str(cell.value)) is not None

            if search1:
                ### FIELD E23 field
                oblik_sheet[f'E{i}'].value = daily_amount
                oblik_obj.save(self.report) # change this field to automatic

            if search2:
                ### FIELD E26 field
                oblik_sheet[f'E{i}'].value = residue_today
                oblik_obj.save(self.report) # change this field to automatic

        print('It\'s ok')
        
    # this date is converted to satisfy needs of LibreOffice date format
    # def convert_date(self):
    #     # date = input('Please select the date of report, use DD-MM-YYYY format: ')
    #     try:
    #         format = '%d-%m-%Y'
    #         convert_to_dt = lambda date: datetime.datetime.strptime(date, format)
    #         date = convert_to_dt(self.date)

    #         # from datetime to date 
    #         date = date.date() ## <class 'datetime.date'>

    #         # get today's date
    #         today = datetime.date.today() ## 2021-07-24 <class 'datetime.date'>
    #         # print(today)
    #         return print(date)
    #     except Exception as err:
    #         print(err)

 
    
    #     ## create a dictionary which allows duplicates keys + nested values
#     #  --> remake to a the LC later
#     items = dict([])

#     for i in range(1, len(reestr_sheet['L'])+1):
#         condition = reestr_sheet[f'L{i}'].value != None
#         if condition:
#             items[reestr_sheet[f'L{i}'].value.lower()] = []

#     ## iterate through reestr sheet cells and get filter by todays date
#     for i in range(1, len(reestr_sheet['A'])+1):
        
#         ## avoid errors when iterating blanc fields
#         condition = reestr_sheet[f'A{i}'].value != None

#         ## use this condition to raise date_error and render to a window of errors
#         ## whnen incorrect date was put or wrong format 
#         condition2 = isinstance(reestr_sheet[f'A{i}'].value, datetime.date) 
        
#         if condition:
#             filter_by_date_cells = list(filter(
#                 # change the date to today if U want to get automatic yesterday report
#                 lambda item: str(item.value)[:10] == converted(date), reestr_sheet[i] 
#                 ))
        
#             if any(filter_by_date_cells):
                
#                 ## this condition will avoid adding Nones to nested dict values
#                 ## avoid: TypeError: unsupported operand type(s) for +: 'int' and 'NoneType'
#                 not_nones_cells = reestr_sheet[f'L{i}'].value != None and reestr_sheet[f'F{i}'].value != None
#                 items[reestr_sheet[f'L{i}'].value.lower()].append(reestr_sheet[f'F{i}'].value) if not_nones_cells else 0
    
#     print(items)


if __name__ == "__main__":
    instance = ParseXLSXData()
    instance.load_file_initial()
    instance.load_file_final()
    instance.clean_fields()
    instance.read_cells()
    instance.add_amounts()
    instance.write_cells()
    instance.input_fuel_arrival()
    instance.date_report_record()
    instance.get_fuel_residue()
    # instance.check_cells()




        # '''
        #     loading initial RT fuel data via openpyxl module from already created MS Excel file
        #     :param name: Describes the comapny name
        #     :type: str max_length = 40
        #     :param country: Depicts the manufacturer's country of origin
        #     :type: str max_length = 20
        #     :param year: depicts the foundation year of a company
        #     :type date: datetime autofiled
        #     :param image: input field for upload/download of images
        #     :type image: 
        #     :param file: input field for upload/download of files
 
        # '''

# def parse_string(*args, **kwargs):
#     '''
#         We will be using all retrieved data from xlsx file
#         as local variables. After slicing the data we would then ask 
#         a user how to save fields or rewrite data to the same/another
#         file. 

#         NOTE: make a script executable file then 
#     '''

#     ## convert and find xlsx file --> make a widget for file location later
#     ## [.....] ADD WIDGET HERE

#     ## Jet A-1 
#     '''
#     registry = 'Реєстр Укртатнафта Jet A-1.xlsx'
#     reestr_lsx_file = Path(
#         path_reserve,
#         'Реєстр Укртатнафта Jet A-1.xlsx'
#         )   
#     '''

#     ## RT 
#     registry = 'Реєстр Укртатнафта РТ.xlsx'
#     reestr_lsx_file = Path(
#         path_reserve,
#         'Реєстр Укртатнафта РТ.xlsx'
#         )

#     reestr_obj = open.load_workbook(reestr_lsx_file)
    
#     '''
#     # Jet A-1
#     report = 'Звіт Jet A-1.xlsx'
#     oblik_lsx_file = Path(
#         path_reserve,
#         'Звіт Jet A-1.xlsx'
#         )
#     '''

#     ## RT
#     report = 'Звіт РТ.xlsx'
#     oblik_lsx_file = Path(
#         path_reserve,
#         report
#         )

#     oblik_obj = open.load_workbook(oblik_lsx_file)

#     # print(type(oblik_obj))

#     reestr_sheet = reestr_obj.active
#     oblik_sheet = oblik_obj.active

#     ## clean fields from a previous day report
#     for i in range(1, len(oblik_sheet['E'])+1):
#         cell = oblik_sheet[f'E{i}']
#         condition = isinstance(cell, Cell)
#         if condition:
#             cell.value = None
#         else:
#             continue
    
#     oblik_obj.save(report)

#     ## convert lambda function for date convert- ---> allows to convert date's if sheet was changed
#     ## NOTE: remake this function after beacuse if the cell is not a timedatae format it'll raise an error
#     ## to an appropriate format: '%Y-%m-%d'
#     converted = lambda date: datetime.date.strftime(date, '%Y-%m-%d')
    
    
#     # get selected date in str format
#     date = input('Please select the date of report, use DD-MM-YYYY format: ')
#     format = '%d-%m-%Y'
#     convert_to_dt = lambda date: datetime.datetime.strptime(date, format)
#     date = convert_to_dt(date)
    
#     # from datetime to date 
#     date = date.date() ## <class 'datetime.date'>
#     # print(dt)
#     # print(type(dt))
    
#     # get today's date
#     today = datetime.date.today() ## 2021-07-24 <class 'datetime.date'>
    
#     # print(today)
#     # print(type(today))
#     # print(dt==today)

#     ## create a dictionary which allows duplicates keys + nested values
#     #  --> remake to a the LC later
#     items = dict([])

#     for i in range(1, len(reestr_sheet['L'])+1):
#         condition = reestr_sheet[f'L{i}'].value != None
#         if condition:
#             items[reestr_sheet[f'L{i}'].value.lower()] = []

#     ## iterate through reestr sheet cells and get filter by todays date
#     for i in range(1, len(reestr_sheet['A'])+1):
        
#         ## avoid errors when iterating blanc fields
#         condition = reestr_sheet[f'A{i}'].value != None

#         ## use this condition to raise date_error and render to a window of errors
#         ## whnen incorrect date was put or wrong format 
#         condition2 = isinstance(reestr_sheet[f'A{i}'].value, datetime.date) 
        
#         if condition:
#             filter_by_date_cells = list(filter(
#                 # change the date to today if U want to get automatic yesterday report
#                 lambda item: str(item.value)[:10] == converted(date), reestr_sheet[i] 
#                 ))
        
#             if any(filter_by_date_cells):
                
#                 ## this condition will avoid adding Nones to nested dict values
#                 ## avoid: TypeError: unsupported operand type(s) for +: 'int' and 'NoneType'
#                 not_nones_cells = reestr_sheet[f'L{i}'].value != None and reestr_sheet[f'F{i}'].value != None
#                 items[reestr_sheet[f'L{i}'].value.lower()].append(reestr_sheet[f'F{i}'].value) if not_nones_cells else 0
    
#     print(items)
    
#     ## getting collected amount values from sheet reestr by carrier
#     ## returns dict with the next data: key - carrier, value - daily amount in kgs
#     def add_amounts(items):
#         daily_amount = 0
#         for key, value in items.items():
#             items[key] = sum(value)
#             daily_amount += sum(value)
#         return items

#     items = add_amounts(items)
#     print(items)

#     ## writing fields E5-E22 using prapared dict of items
#     ## with 'Zvit' Fields C5-C22 as keys
#     ## stage2: write matched by key values into cells through
#     ## get from loop indices - i
#     try:

#         i = 0
#         for carrier_cell in oblik_sheet[f'C']:

#             ## condition willl avoid exceptions
#             ##  when using blank cell as a key 
#             condition4 = carrier_cell.value is not None
#             record_value = items.get(str(carrier_cell.value).lower()) if condition4 else 0
            
#             ## iteration should be in this block 
#             ## as we need to get all indices of cells and sort only those
#             ## where carrier/key - value from items were found
#             i += 1
#             if record_value:
#                 oblik_sheet[f'E{i}'].value = record_value            
#                 oblik_obj.save(report) # change this field to automatic     

#     except:
#         print('Render a possible error here to the window')

    
#     ## input fuel arrival:
#     try:
#         fuel_arrived = int(input('Введіть надходження палива, кг: '))
#         i = 0
#         for cell in oblik_sheet[f'C']:
#             i += 1
#             pattern = re.compile('Прибуток палива')
#             ## find appropriate cell condition to write daily fuel suppply 
#             result = pattern.match(str(cell.value)) is not None
#             if result:
#                 oblik_sheet[f'E{i}'].value = fuel_arrived 
#                 oblik_obj.save(report) # change this field to automatic

#     except ValueError as input_err:
#         print('Please do not enter letters')
#         pass

#     except (TypeError, AttributeError) as input_error:
#         print('Please enter a correct amount')
#         pass


    
#     ## date of he report header B3
#     try:
#         ## block for automatic yesterday report
#         # automatic date report 
#         # yest = datetime.date.today() - datetime.timedelta(days=1)
#         # yest = datetime.date.strftime(yest, '%d.%m.%Y')
#         # today = datetime.date.strftime(today, '%d.%m.%Y')

#         ## variable value will be added for todays residue calculation
#         # oblik_sheet[f'B{3}'].value = f'Дата: {today}'
#         # oblik_obj.save(report) ## change this field to automatic
#         # #print(fuel_residue) if condition else 'Nothing'
        
#         ## block for preselected date of the report
#         ## date for prior of this report date
#         prior_rep_date = date - datetime.timedelta(days=1)
#         prior_rep_date = datetime.date.strftime(prior_rep_date, '%d.%m.%Y')
#         date = datetime.date.strftime(date, '%d.%m.%Y')
        
#         oblik_sheet[f'B{3}'].value = f'Дата: {date}'
#         oblik_obj.save(report) ## change this field to automatic
        
#     except (TypeError, ValueError) as input_error:
#         print(input_error)
#         pass

#     ## input fuel residue:
#     ## up-to-date fuel residue calculation (at the end of a current date 23:59:59)
#     try:

#         fuel_residue = int(input(f'Введіть залишок палива КГ, на дату {prior_rep_date} 23:59: '))
#         ## avoid errors when blanc field of fuel supply (no fuel supply daily)
#         condition = fuel_residue is not None
#         # E23
#         daily_amount = sum([i for i in items.values()])
        
#         # E26
#         residue_today = fuel_residue + fuel_arrived - daily_amount
#         pattern1 = re.compile('Всього видано на пероні')
#         pattern2 = re.compile('Залишок на складі ПММ')

#         i = 0
#         for cell in oblik_sheet[f'C']:
#             i += 1
#             search1 = pattern1.match(str(cell.value)) is not None
#             search2 = pattern2.match(str(cell.value)) is not None

#             if search1:
#                 ### FIELD E23 field
#                 oblik_sheet[f'E{i}'].value = daily_amount
#                 oblik_obj.save(report) # change this field to automatic

#             if search2:
#                 ### FIELD E26 field
#                 oblik_sheet[f'E{i}'].value = residue_today
#                 oblik_obj.save(report) # change this field to automatic
#     except Exception as err:
#         print(err)
#         pass 

            
# parse_string()